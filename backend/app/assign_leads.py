"""Add categories, create reps, and assign leads equally per category."""
import sqlite3
import openpyxl
import asyncio

DB = r"C:\Users\cloud\OneDrive\Desktop\Hybrid_Second_Brain\agency dashboard\backend\sales_dashboard.db"
XLSX = r"C:\Users\cloud\OneDrive\Desktop\Hybrid_Second_Brain\03_Active_Projects\bots\leads-scraper\output\Kochi_Leads_Enriched.xlsx"


def backfill_categories():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(leads)")
    cols = [c[1] for c in cur.fetchall()]
    if "category" not in cols:
        cur.execute("ALTER TABLE leads ADD COLUMN category TEXT")
        print("Added category column")

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb.active
    raw_headers = [str(h).strip() if h else "" for h in next(ws.iter_rows(values_only=True))]
    headers = [h.lower().replace(" ", "_") for h in raw_headers]

    leads_data = []
    for row in ws.iter_rows(values_only=True):
        data = {}
        for i, val in enumerate(row):
            if i < len(headers):
                data[headers[i]] = str(val).strip() if val else ""
        bn = data.get("business_name", "").strip()
        ph = data.get("phone", "").strip()
        cat = data.get("category", "").strip()
        if bn and ph and cat and cat.lower() != "category":
            leads_data.append((bn, ph, cat))

    print(f"Loaded {len(leads_data)} leads from XLSX with categories")

    updated = 0
    for bn, ph, cat in leads_data:
        cur.execute(
            "UPDATE leads SET category = ? WHERE business_name = ? AND phone = ? AND (category IS NULL OR category = '')",
            (cat, bn, ph),
        )
        if cur.rowcount > 0:
            updated += cur.rowcount

    conn.commit()

    cur.execute("SELECT COUNT(*) FROM leads WHERE category IS NOT NULL AND category != ''")
    with_cat = cur.fetchone()[0]
    print(f"Leads with category: {with_cat}, Updated: {updated}")

    wb.close()
    return conn


def create_reps_and_assign(conn):
    import sys
    sys.path.insert(0, r"C:\Users\cloud\OneDrive\Desktop\Hybrid_Second_Brain\agency dashboard\backend")

    from app.database import async_session_factory, init_db
    from app.models.user import User, UserRole
    from app.services.auth import hash_password
    from sqlalchemy import select

    async def _create():
        await init_db()
        async with async_session_factory() as db:
            reps_info = [
                ("ashba@agency.com", "ashba123", "Ashba"),
                ("jai@agency.com", "jai123", "Jai"),
                ("agnas@agency.com", "agnas123", "Agnas"),
            ]
            created = []
            for email, pw, name in reps_info:
                result = await db.execute(select(User).where(User.email == email))
                if not result.scalar_one_or_none():
                    user = User(
                        email=email,
                        hashed_password=hash_password(pw),
                        full_name=name,
                        role=UserRole.REP,
                    )
                    db.add(user)
                    created.append(name)
            await db.commit()
            print(f"Created rep users: {created}")

    asyncio.run(_create())

    conn = conn_override or sqlite3.connect(DB)
    cur = conn.cursor()

    cur.execute("SELECT id, full_name FROM users WHERE role = 'rep' ORDER BY id")
    reps = cur.fetchall()
    print(f"Reps: {reps}")

    cur.execute(
        "SELECT DISTINCT category FROM leads WHERE category IS NOT NULL AND category != '' ORDER BY category"
    )
    categories = [c[0] for c in cur.fetchall()]
    print(f"Categories: {len(categories)}")

    assignments = {r[0]: [] for r in reps}

    for cat in categories:
        cur.execute(
            "SELECT id FROM leads WHERE category = ? AND assigned_to IS NULL ORDER BY id",
            (cat,),
        )
        lead_ids = [r[0] for r in cur.fetchall()]
        if not lead_ids:
            continue
        base = len(lead_ids) // len(reps)
        rem = len(lead_ids) % len(reps)
        idx = 0
        for i, rep in enumerate(reps):
            count = base + (1 if i < rem else 0)
            chunk = lead_ids[idx : idx + count]
            idx += count
            for lid in chunk:
                cur.execute("UPDATE leads SET assigned_to = ? WHERE id = ?", (rep[0], lid))
                assignments[rep[0]].append(lid)

    conn.commit()

    for rid, lids in assignments.items():
        name = next(r[1] for r in reps if r[0] == rid)
        print(f"  {name}: {len(lids)} leads")

    conn.close()


if __name__ == "__main__":
    conn = backfill_categories()
    create_reps_and_assign(conn)
    print("Done!")
