import csv
from typing import Iterator
from openpyxl import load_workbook


def parse_lead_file(file_path: str) -> Iterator[dict]:
    if file_path.endswith(".csv"):
        yield from _parse_csv(file_path)
    elif file_path.endswith((".xls", ".xlsx")):
        yield from _parse_xlsx(file_path)
    else:
        raise ValueError("Unsupported file format. Use CSV or Excel.")


def _normalize_headers(headers: list[str]) -> list[str]:
    h2 = []
    for h in headers:
        h = h.strip().lower().replace(" ", "_")
        alias = {
            "business_name": "business_name",
            "company_name": "business_name",
            "company": "business_name",
            "business": "business_name",
            "name": "business_name",
            "contact_name": "contact_name",
            "contact_person": "contact_name",
            "contact": "contact_name",
            "phone": "phone",
            "telephone": "phone",
            "phone_number": "phone",
            "mobile": "phone",
            "email": "email",
            "e_mail": "email",
            "website": "website",
            "web": "website",
            "url": "website",
            "address": "address",
            "category": "category",
            "area": "area",
            "rating": "rating",
            "reviews": "reviews",
            "need_score": "need_score",
        }
        h2.append(alias.get(h, h))
    return h2


def _parse_csv(file_path: str) -> Iterator[dict]:
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("Empty CSV file")
        normalized = _normalize_headers(reader.fieldnames)
        if "business_name" not in normalized or "phone" not in normalized:
            raise ValueError("Missing required columns: business_name and phone")

        for row in reader:
            data = {}
            for raw_key, norm_key in zip(reader.fieldnames, normalized):
                data[norm_key] = row[raw_key].strip() if row.get(raw_key) else None
            yield {
                "business_name": data.get("business_name", ""),
                "contact_name": data.get("contact_name") or data.get("business_name", ""),
                "phone": data.get("phone", ""),
                "email": data.get("email"),
                "website": data.get("website"),
            }


def _parse_xlsx(file_path: str) -> Iterator[dict]:
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = iter(ws.iter_rows(values_only=True))
    raw_headers = [str(h).strip() if h else "" for h in next(rows, [])]
    headers = _normalize_headers(raw_headers)

    if "business_name" not in headers or "phone" not in headers:
        raise ValueError("Missing required columns: business_name and phone")

    for row in rows:
        data = {}
        for i, val in enumerate(row):
            if i < len(headers):
                data[headers[i]] = str(val).strip() if val else None
        if data.get("business_name") and data.get("phone"):
            yield {
                "business_name": data.get("business_name", ""),
                "contact_name": data.get("contact_name") or data.get("business_name", ""),
                "phone": data.get("phone", ""),
                "email": data.get("email"),
                "website": data.get("website"),
            }
    wb.close()
